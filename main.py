import tkinter as tk
from tkinter import ttk, filedialog
import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import openpyxl
import os
from pyzbar.pyzbar import decode
from PIL import Image
import cv2

class Inventory:
    def __init__(self):
        self.products = {}

    def add_product(self, name, quantity, price):
        date = datetime.date.today()
        if name in self.products:
            self.products[name]['quantity'] += quantity
        else:
            self.products[name] = {'quantity': quantity, 'price': price, 'date': date}

    def remove_product(self, name, quantity):
        if name in self.products and self.products[name]['quantity'] >= quantity:
            self.products[name]['quantity'] -= quantity
            if self.products[name]['quantity'] == 0:
                del self.products[name]
            return True
        else:
            return False

    def view_inventory(self):
        if not self.products:
            return "Inventory is empty."
        else:
            result = ""
            for name, details in self.products.items():
                result += f"{name}: Quantity - {details['quantity']}, Price - {details['price']}\n"
            return result

    def get_total_value(self):
        total = sum(details['quantity'] * details['price'] for details in self.products.values())
        return f"Total inventory value: {total}"

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.inventory = Inventory()
        self.title("Inventory Management System for Gudacas Litsong Manok")

        # Input buttons frame
        input_button_frame = tk.Frame(self)
        input_button_frame.pack(pady=10)
        self.input_button = tk.Button(input_button_frame, text="Input Products", command=self.open_input_popup)
        self.input_button.grid(row=0, column=0, padx=5)
        self.qr_button = tk.Button(input_button_frame, text="Scan QR Code", command=self.open_scan_popup)
        self.qr_button.grid(row=0, column=1, padx=5)

        # Buttons frame
        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)
        tk.Button(button_frame, text="Save Report", command=self.save_report).grid(row=0, column=0, padx=5)
        tk.Button(button_frame, text="Remove Selected", command=self.remove_selected).grid(row=0, column=1, padx=5)

        # Inventory frame
        inventory_frame = tk.Frame(self)
        inventory_frame.pack(pady=10, fill=tk.BOTH, expand=True)
        tk.Label(inventory_frame, text="Inventory:").pack()
        # Search
        search_frame = tk.Frame(inventory_frame)
        search_frame.pack(fill=tk.X)
        tk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var)
        self.search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_entry.bind("<KeyRelease>", self.filter_inventory)
        self.tree = ttk.Treeview(inventory_frame, columns=("Name", "Quantity", "Unit Price", "Total Price", "Date"), show="headings", height=15)
        self.tree.heading("Name", text="Name")
        self.tree.heading("Quantity", text="Quantity")
        self.tree.heading("Unit Price", text="Unit Price")
        self.tree.heading("Total Price", text="Total Price")
        self.tree.heading("Date", text="Date")
        self.tree.column("Name", anchor='center')
        self.tree.column("Quantity", anchor='center')
        self.tree.column("Unit Price", anchor='center')
        self.tree.column("Total Price", anchor='center')
        self.tree.column("Date", anchor='center')
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Messages frame
        messages_frame = tk.Frame(self)
        messages_frame.pack(pady=10, fill=tk.X)
        tk.Label(messages_frame, text="Messages:").pack()
        self.output = tk.Text(messages_frame, height=3, width=60)
        self.output.pack(fill=tk.X)

    def add_product(self):
        try:
            name = self.name_entry.get()
            qty = int(self.qty_entry.get())
            price = float(self.price_entry.get())
            self.inventory.add_product(name, qty, price)
            self.output.insert(tk.END, f"Added {name}\n")
            self.clear_entries()
        except ValueError:
            self.output.insert(tk.END, "Invalid input for add product.\n")

    def remove_product(self):
        try:
            name = self.remove_name_entry.get()
            qty = int(self.remove_qty_entry.get())
            if self.inventory.remove_product(name, qty):
                self.output.insert(tk.END, f"Removed {qty} of {name}\n")
            else:
                self.output.insert(tk.END, f"Insufficient quantity for {name}\n")
            self.clear_remove_entries()
        except ValueError:
            self.output.insert(tk.END, "Invalid input for remove product.\n")

    def filter_inventory(self, event=None):
        search_term = self.search_var.get().lower()
        # Clear the tree
        for item in self.tree.get_children():
            self.tree.delete(item)
        # Insert filtered items
        for name, details in self.inventory.products.items():
            if search_term in name.lower():
                total_price = details['quantity'] * details['price']
                self.tree.insert("", tk.END, values=(name, details['quantity'], details['price'], total_price, str(details['date'])))

    def view_inventory(self):
        self.filter_inventory()

    def save_report(self):
        filename = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF files", "*.pdf"), ("Excel files", "*.xlsx")])
        if not filename:
            return
        if filename.endswith('.pdf'):
            self.save_as_pdf(filename)
        elif filename.endswith('.xlsx'):
            self.save_as_excel(filename)
        else:
            self.output.insert(tk.END, "Unsupported file type.\n")

    def save_as_pdf(self, filename):
        try:
            doc = SimpleDocTemplate(filename, pagesize=letter)
            elements = []
            data = [["Name", "Quantity", "Unit Price", "Total Price", "Date"]]
            for name, details in self.inventory.products.items():
                total_price = details['quantity'] * details['price']
                data.append([name, str(details['quantity']), str(details['price']), str(total_price), str(details['date'])])
            if len(data) == 1:
                data.append(["No products", "", "", "", ""])
            table = Table(data)
            style = TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.grey),
                ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 14),
                ('BOTTOMPADDING', (0,0), (-1,0), 12),
                ('BACKGROUND', (0,1), (-1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black),
            ])
            table.setStyle(style)
            elements.append(table)
            doc.build(elements)
            self.output.insert(tk.END, f"PDF report saved to {filename}\n")
        except Exception as e:
            self.output.insert(tk.END, f"Error saving PDF: {str(e)}\n")

    def save_as_excel(self, filename):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory Report"
            headers = ["Name", "Quantity", "Unit Price", "Total Price", "Date"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            row = 2
            for name, details in self.inventory.products.items():
                total_price = details['quantity'] * details['price']
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=details['quantity'])
                ws.cell(row=row, column=3, value=details['price'])
                ws.cell(row=row, column=4, value=total_price)
                ws.cell(row=row, column=5, value=str(details['date']))
                row += 1
            if row == 2:
                ws.cell(row=2, column=1, value="No products")
            wb.save(filename)
            self.output.insert(tk.END, f"Excel report saved to {filename}\n")
        except Exception as e:
            self.output.insert(tk.END, f"Error saving Excel: {str(e)}\n")

    def open_scan_popup(self):
        popup = tk.Toplevel(self)
        popup.title("Scan QR Code")
        popup.geometry("300x150")
        tk.Label(popup, text="Select Scan Method", font=("Arial", 12, "bold")).pack(pady=10)
        tk.Button(popup, text="Upload Image", command=lambda: [self.scan_qr_code(), popup.destroy()]).pack(pady=5)
        tk.Button(popup, text="Camera Scan", command=lambda: [self.camera_scan_qr(), popup.destroy()]).pack(pady=5)

    def camera_scan_qr(self):
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            self.output.insert(tk.END, "Camera not available.\n")
            return
        self.output.insert(tk.END, "Camera scanning... Point QR code at camera.\n")
        found_qr = False
        max_frames = 500
        frame_count = 0
        while not found_qr and frame_count < max_frames:
            ret, frame = cap.read()
            if not ret:
                break
            decoded_objects = decode(frame)
            if decoded_objects:
                qr_data = decoded_objects[0].data.decode('utf-8')
                self.output.insert(tk.END, f"QR Code found: {qr_data}\n")
                found_qr = True
                self.open_qr_input_popup(qr_data)
            frame_count += 1
        cap.release()
        if not found_qr:
            self.output.insert(tk.END, "No QR code scanned.\n")

    def scan_qr_code(self):
        qr_file = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg *.png *.bmp")])
        if not qr_file:
            return
        try:
            image = Image.open(qr_file)
            decoded_objects = decode(image)
            if decoded_objects:
                qr_data = decoded_objects[0].data.decode('utf-8')
                self.output.insert(tk.END, f"QR Code scanned: {qr_data}\n")
                self.open_input_popup(qr_data)
            else:
                self.output.insert(tk.END, "No QR code found in image.\n")
        except Exception as e:
            self.output.insert(tk.END, f"Error scanning QR: {str(e)}\n")

    def open_qr_input_popup(self, product_name):
        popup = tk.Toplevel(self)
        popup.title("Add Product from QR Code")
        popup.geometry("300x200")
        tk.Label(popup, text="Product Details", font=("Arial", 12, "bold")).pack(pady=10)
        tk.Label(popup, text=f"Product: {product_name}").pack()
        tk.Label(popup, text="Quantity:").pack()
        qty_entry = tk.Entry(popup)
        qty_entry.pack()
        tk.Label(popup, text="Unit Price:").pack()
        price_entry = tk.Entry(popup)
        price_entry.pack()
        tk.Button(popup, text="Add Product", command=lambda: self.add_qr_product(product_name, qty_entry, price_entry, popup)).pack(pady=10)

    def add_qr_product(self, product_name, qty_entry, price_entry, popup):
        try:
            qty = int(qty_entry.get())
            price = float(price_entry.get())
            self.inventory.add_product(product_name, qty, price)
            self.output.insert(tk.END, f"Added {product_name} (Qty: {qty}, Price: {price})\n")
            self.search_var.set("")
            self.view_inventory()
            popup.destroy()
        except ValueError:
            self.output.insert(tk.END, "Invalid quantity or price.\n")

    def remove_selected(self):
        selected = self.tree.selection()
        if not selected:
            self.output.insert(tk.END, "No item selected.\n")
            return
        item = self.tree.item(selected[0])
        name = item['values'][0]
        if name not in self.inventory.products:
            self.output.insert(tk.END, "Item not found.\n")
            return
        from tkinter import messagebox
        if messagebox.askyesno("Confirm Remove", f"Are you sure you want to remove '{name}'?"):
            qty = self.inventory.products[name]['quantity']
            if self.inventory.remove_product(name, qty):
                self.output.insert(tk.END, f"Removed {name}\n")
                self.search_var.set("")
                self.view_inventory()
            else:
                self.output.insert(tk.END, f"Error removing {name}\n")

    def add_from_popup(self, name_entry, qty_entry, price_entry, popup):
        try:
            name = name_entry.get()
            qty = int(qty_entry.get())
            price = float(price_entry.get())
            self.inventory.add_product(name, qty, price)
            self.output.insert(tk.END, f"Added {name}\n")
            self.search_var.set("")
            self.view_inventory()
            popup.destroy()
        except ValueError:
            self.output.insert(tk.END, "Invalid input for add product.\n")

    def open_input_popup(self, qr_data=None):
        popup = tk.Toplevel(self)
        popup.title("Manage Products")
        popup.geometry("300x250")

        # Add product frame in popup
        add_frame = tk.Frame(popup)
        add_frame.pack(pady=10)
        tk.Label(add_frame, text="Add Product", font=("Arial", 12, "bold")).grid(row=0, column=0, columnspan=2)
        tk.Label(add_frame, text="Name:").grid(row=1, column=0, sticky="e")
        name_entry = tk.Entry(add_frame)
        name_entry.grid(row=1, column=1)
        if qr_data:
            name_entry.insert(0, qr_data)
        tk.Label(add_frame, text="Quantity:").grid(row=2, column=0, sticky="e")
        qty_entry = tk.Entry(add_frame)
        qty_entry.grid(row=2, column=1)
        tk.Label(add_frame, text="Price:").grid(row=3, column=0, sticky="e")
        price_entry = tk.Entry(add_frame)
        price_entry.grid(row=3, column=1)
        tk.Button(add_frame, text="Add", command=lambda: self.add_from_popup(name_entry, qty_entry, price_entry, popup)).grid(row=4, column=0, columnspan=2)

if __name__ == "__main__":
    app = App()
    app.mainloop()